# validation_logic.py
# Complete validation logic extracted from the original Streamlit code

import pandas as pd
import json
import re
import os
from datetime import datetime
from typing import Dict, List, Any, Tuple, Optional, Union
from dataclasses import dataclass, asdict
from abc import ABC, abstractmethod
import logging
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import tempfile
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class ValidationRule:
    """Enhanced validation rule for Excel workbench format"""
    rule_id: str
    target_sheet: str
    cell_column_range: str
    condition: str
    error_message: str
    severity: str = "error"
    priority: int = 1
    active: bool = True

@dataclass
class ValidationViolation:
    """Represents a validation violation"""
    rule_id: str
    sheet_name: str
    cell_address: str
    row_index: int
    column_name: str
    value: Any
    message: str
    severity: str
    suggested_fix: Optional[str] = None

class RuleInterpreterAgent:
    """Agent responsible for interpreting and executing validation rules"""
    
    def __init__(self):
        self.supported_operators = {
            '>': lambda x, y: float(x) > float(y),
            '>=': lambda x, y: float(x) >= float(y),
            '<': lambda x, y: float(x) < float(y),
            '<=': lambda x, y: float(x) <= float(y),
            '==': lambda x, y: str(x) == str(y),
            '!=': lambda x, y: str(x) != str(y),
            'contains': lambda x, y: str(y).lower() in str(x).lower(),
            'not_contains': lambda x, y: str(y).lower() not in str(x).lower(),
            'starts_with': lambda x, y: str(x).lower().startswith(str(y).lower()),
            'ends_with': lambda x, y: str(x).lower().endswith(str(y).lower()),
        }
    
    def parse_range(self, range_str: str) -> Dict[str, Any]:
        """Parse Excel range string into actionable components"""
        range_str = range_str.strip().upper()
        
        if ':' in range_str:
            start, end = range_str.split(':')
            if start.isalpha() and end.isalpha():
                return {
                    'type': 'column_range',
                    'start_col': start,
                    'end_col': end,
                    'columns': [start] if start == end else list(range(ord(start), ord(end) + 1))
                }
            else:
                return {
                    'type': 'cell_range',
                    'start_cell': start,
                    'end_cell': end,
                    'range': range_str
                }
        else:
            if range_str.isalpha():
                return {
                    'type': 'single_column',
                    'column': range_str
                }
            else:
                return {
                    'type': 'single_cell',
                    'cell': range_str
                }
    
    def get_column_letter_from_index(self, index: int) -> str:
        """Convert column index to letter"""
        return get_column_letter(index)
    
    def get_column_index_from_letter(self, letter: str) -> int:
        """Convert column letter to index"""
        return column_index_from_string(letter)
    
    def _parse_accounting_value(self, value: str) -> str:
        """Convert accounting format (5) to -5"""
        if isinstance(value, str):
            if re.match(r'^\(.*\)$', value.strip()):
                return '-' + value.strip()[1:-1]
        return value
    
    def _convert_to_numeric(self, value: Any) -> float:
        """Convert value to numeric, handling various formats including accounting"""
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            value = self._parse_accounting_value(value)
            cleaned = value.replace(',', '').replace('$', '').replace('%', '').strip()
            
            if cleaned == '' or cleaned.lower() in ['na', 'null', 'none', '-']:
                raise ValueError("Cannot convert to numeric")
            
            if '(' in cleaned and ')' in cleaned:
                cleaned = cleaned.replace('(', '').replace(')', '')
                if not cleaned.startswith('-'):
                    cleaned = '-' + cleaned
            
            return float(cleaned)
        
        return float(value)
    
    def interpret_rule(self, rule: ValidationRule, value: Any) -> bool:
        """
        Interpret and execute a validation rule against a value
        Returns True if VIOLATION (condition NOT met)
        Returns False if OK (condition IS met)
        """
        try:
            if pd.isna(value) or value == '' or value is None:
                condition = rule.condition.strip().lower()
                if 'not_empty' in condition:
                    return True
                else:
                    return True
            
            condition = rule.condition.strip()
            
            if condition.lower().startswith('regex:'):
                pattern = condition[6:].strip()
                is_match = bool(re.match(pattern, str(value)))
                return not is_match
            
            elif condition.lower().startswith('date_future'):
                try:
                    date_val = pd.to_datetime(value)
                    is_future = date_val > datetime.now()
                    return not is_future
                except:
                    return True
            
            elif condition.lower().startswith('date_past'):
                try:
                    date_val = pd.to_datetime(value)
                    is_past = date_val < datetime.now()
                    return not is_past
                except:
                    return True
            
            elif condition.lower().startswith('not_empty'):
                is_not_empty = value != '' and not pd.isna(value)
                return not is_not_empty
            
            elif condition.lower().startswith('is_number'):
                try:
                    self._convert_to_numeric(value)
                    return False
                except:
                    return True
            
            else:
                for op_str, op_func in self.supported_operators.items():
                    if condition.startswith(op_str):
                        threshold = condition.replace(op_str, '').strip()
                        
                        if '%' in threshold:
                            threshold = float(threshold.replace('%', '')) / 100
                        
                        try:
                            if op_str in ['>', '>=', '<', '<=']:
                                numeric_value = self._convert_to_numeric(value)
                                numeric_threshold = float(threshold)
                                
                                if op_str == '>':
                                    condition_met = numeric_value > numeric_threshold
                                elif op_str == '>=':
                                    condition_met = numeric_value >= numeric_threshold
                                elif op_str == '<':
                                    condition_met = numeric_value < numeric_threshold
                                elif op_str == '<=':
                                    condition_met = numeric_value <= numeric_threshold
                                
                                return not condition_met
                            else:
                                condition_met = op_func(value, threshold)
                                return not condition_met
                                
                        except (ValueError, TypeError):
                            return True
            
            return True
            
        except Exception as e:
            logger.error(f"Error interpreting rule {rule.rule_id}: {e}")
            return True

class SmartRuleInterpreter:
    """Extended rule interpreter that handles complex validation patterns"""
    
    def __init__(self, base_interpreter):
        self.base_interpreter = base_interpreter
        self.smart_rule_handlers = {
            'gap_detection': self._handle_gap_detection,
            'conditional_comparison': self._handle_conditional_comparison,
            'conditional_sum': self._handle_conditional_sum,
            'tnt_bill_detection': self._handle_tnt_bill_detection,
        }
    
    def _format_number(self, value: Any) -> str:
        """Format numbers for display with reasonable rounding"""
        try:
            if pd.isna(value) or value == '' or value is None:
                return str(value)
            
            num_value = float(value)
            
            if abs(num_value) < 0.01:
                return "0"
            
            formatted = f"{num_value:,.2f}".rstrip('0').rstrip('.')
            return formatted
            
        except (ValueError, TypeError):
            return str(value)
    
    def interpret_smart_rule(self, rule: ValidationRule, df: pd.DataFrame, workbook_data: Dict = None) -> List[ValidationViolation]:
        """Handle complex multi-cell validation rules"""
        violations = []
        condition = rule.condition.strip().lower()
        
        for rule_type, handler in self.smart_rule_handlers.items():
            if condition.startswith(rule_type):
                try:
                    violations = handler(rule, df, condition, workbook_data)
                except Exception as e:
                    logger.error(f"Error in smart rule {rule.rule_id}: {e}")
                    violations = [ValidationViolation(
                        rule_id=rule.rule_id,
                        sheet_name=rule.target_sheet,
                        cell_address="Rule Error",
                        row_index=0,
                        column_name="System",
                        value="Rule execution failed",
                        message=f"Rule execution error: {e}",
                        severity="error"
                    )]
                break
        
        return violations
    
    def _handle_gap_detection(self, rule: ValidationRule, df: pd.DataFrame, condition: str, workbook_data: Dict) -> List[ValidationViolation]:
        """Handle gap detection: gap_detection:start_col:end_col:start_row:end_row"""
        violations = []
        
        try:
            parts = condition.split(':')
            if len(parts) != 5:
                raise ValueError("Gap detection requires format: gap_detection:start_col:end_col:start_row:end_row")
            
            start_col = parts[1].upper()
            end_col = parts[2].upper() 
            start_row = int(parts[3])
            end_row = int(parts[4])
            
            start_row_idx = start_row - 2
            end_row_idx = end_row - 2
            start_col_idx = self.base_interpreter.get_column_index_from_letter(start_col) - 1
            end_col_idx = self.base_interpreter.get_column_index_from_letter(end_col) - 1
            
            # Find actual start and end columns with data
            actual_start_col_idx = None
            for col_idx in range(start_col_idx, min(end_col_idx + 1, len(df.columns))):
                if col_idx >= len(df.columns):
                    continue
                    
                has_data = False
                for row_idx in range(max(0, start_row_idx), min(end_row_idx + 1, len(df))):
                    value = df.iloc[row_idx, col_idx]
                    if pd.notna(value) and str(value).strip() != '':
                        has_data = True
                        break
                
                if has_data:
                    actual_start_col_idx = col_idx
                    break
            
            actual_end_col_idx = None
            for col_idx in range(min(end_col_idx, len(df.columns) - 1), start_col_idx - 1, -1):
                if col_idx >= len(df.columns):
                    continue
                    
                has_data = False
                for row_idx in range(max(0, start_row_idx), min(end_row_idx + 1, len(df))):
                    value = df.iloc[row_idx, col_idx]
                    if pd.notna(value) and str(value).strip() != '':
                        has_data = True
                        break
                
                if has_data:
                    actual_end_col_idx = col_idx
                    break
            
            if actual_start_col_idx is None or actual_end_col_idx is None:
                return violations
            
            start_excel_letter = self.base_interpreter.get_column_letter_from_index(actual_start_col_idx + 1)
            end_excel_letter = self.base_interpreter.get_column_letter_from_index(actual_end_col_idx + 1)
            
            # Check for completely empty columns between start and end
            completely_empty_columns = []
            
            for col_idx in range(actual_start_col_idx + 1, actual_end_col_idx):
                if col_idx >= len(df.columns):
                    continue
                    
                has_any_data = False
                for row_idx in range(max(0, start_row_idx), min(end_row_idx + 1, len(df))):
                    value = df.iloc[row_idx, col_idx]
                    if pd.notna(value) and str(value).strip() != '':
                        has_any_data = True
                        break
                
                if not has_any_data:
                    excel_col_letter = self.base_interpreter.get_column_letter_from_index(col_idx + 1)
                    completely_empty_columns.append(excel_col_letter)
            
            # Create violations for empty columns
            for excel_col in completely_empty_columns:
                clean_message = f"{rule.error_message}\n\nStart: Column {start_excel_letter}\nEnd: Column {end_excel_letter}"
                
                violations.append(ValidationViolation(
                    rule_id=rule.rule_id,
                    sheet_name=rule.target_sheet,
                    cell_address=f"Column {excel_col}",
                    row_index=0,
                    column_name=f"Column {excel_col}",
                    value="Missing",
                    message=clean_message,
                    severity=rule.severity,
                    suggested_fix=f"Add data to Column {excel_col}"
                ))
        
        except Exception as e:
            logger.error(f"Gap detection error: {e}")
            raise
        
        return violations

    def _handle_conditional_comparison(self, rule: ValidationRule, df: pd.DataFrame, condition: str, workbook_data: Dict) -> List[ValidationViolation]:
        """Handle conditional comparison: conditional_comparison:cell1:cell2:condition_type"""
        violations = []
        
        try:
            parts = condition.split(':')
            if len(parts) != 4:
                raise ValueError("Conditional comparison requires format: conditional_comparison:cell1:cell2:condition_type")
            
            cell1_addr = parts[1].upper()
            cell2_addr = parts[2].upper()
            condition_type = parts[3].lower()
            
            cell1_value = self._get_cell_value(df, cell1_addr)
            cell2_value = self._get_cell_value(df, cell2_addr)
            
            cell1_name = self._get_cell_display_name(cell1_addr)
            cell2_name = self._get_cell_display_name(cell2_addr)
            
            violation_found = False
            violation_message = rule.error_message
            suggested_fix = ""
            
            if condition_type == "greater_or_zero":
                if pd.notna(cell1_value) and str(cell1_value).strip() != '':
                    try:
                        cell1_num = float(cell1_value)
                        cell2_num = float(cell2_value) if pd.notna(cell2_value) else 0
                        
                        if cell2_num == 0 or cell1_num > cell2_num:
                            violation_found = True
                            violation_message = f"{rule.error_message} ({cell1_name}={self._format_number(cell1_value)}, {cell2_name}={self._format_number(cell2_value)})"
                            suggested_fix = f"Ensure {cell2_name} has appropriate value relative to {cell1_name}"
                    except (ValueError, TypeError):
                        pass
            
            if violation_found:
                violations.append(ValidationViolation(
                    rule_id=rule.rule_id,
                    sheet_name=rule.target_sheet,
                    cell_address=f"{cell1_addr},{cell2_addr}",
                    row_index=0,
                    column_name="Conditional",
                    value=f"{cell1_name}={self._format_number(cell1_value)}, {cell2_name}={self._format_number(cell2_value)}",
                    message=violation_message,
                    severity=rule.severity,
                    suggested_fix=suggested_fix
                ))
        
        except Exception as e:
            logger.error(f"Conditional comparison error: {e}")
            raise
        
        return violations
    
    def _handle_conditional_sum(self, rule: ValidationRule, df: pd.DataFrame, condition: str, workbook_data: Dict) -> List[ValidationViolation]:
        """Handle conditional sum: conditional_sum:lookup_col:lookup_value:sum_col:target_cell"""
        violations = []
        
        try:
            parts = condition.split(':')
            if len(parts) != 5:
                raise ValueError("Conditional sum requires format: conditional_sum:lookup_col:lookup_value:sum_col:target_cell")
            
            lookup_col = parts[1].upper()
            lookup_value = parts[2]
            sum_col = parts[3].upper()
            target_cell = parts[4].upper()
            
            target_value = self._get_cell_value(df, target_cell)
            target_name = self._get_cell_display_name(target_cell)
            
            lookup_col_idx = self.base_interpreter.get_column_index_from_letter(lookup_col) - 1
            sum_col_idx = self.base_interpreter.get_column_index_from_letter(sum_col) - 1
            
            matching_sum = 0
            matching_rows = []
            
            if lookup_col_idx < len(df.columns) and sum_col_idx < len(df.columns):
                lookup_column_name = df.columns[lookup_col_idx]
                sum_column_name = df.columns[sum_col_idx]
                
                for idx, row in df.iterrows():
                    lookup_cell_value = row[lookup_column_name]
                    if pd.notna(lookup_cell_value) and lookup_value.lower() in str(lookup_cell_value).lower():
                        sum_cell_value = row[sum_column_name]
                        try:
                            if pd.notna(sum_cell_value) and str(sum_cell_value).strip() != '':
                                matching_sum += float(sum_cell_value)
                                matching_rows.append(idx + 2)
                        except (ValueError, TypeError):
                            pass
            
            # Compare with target
            try:
                target_num = float(target_value) if pd.notna(target_value) else 0
                if abs(matching_sum - target_num) > 0.01:
                    violations.append(ValidationViolation(
                        rule_id=rule.rule_id,
                        sheet_name=rule.target_sheet,
                        cell_address=target_cell,
                        row_index=0,
                        column_name="Sum Validation",
                        value=f"{target_name}: {self._format_number(target_num)}, Calculated: {self._format_number(matching_sum)}",
                        message=f"{rule.error_message} (Expected {target_name}: {self._format_number(target_num)}, Got: {self._format_number(matching_sum)})",
                        severity=rule.severity,
                        suggested_fix=f"Check {lookup_value} entries in rows {matching_rows} or update {target_name}"
                    ))
            except (ValueError, TypeError):
                violations.append(ValidationViolation(
                    rule_id=rule.rule_id,
                    sheet_name=rule.target_sheet,
                    cell_address=target_cell,
                    row_index=0,
                    column_name="Sum Validation",
                    value=str(target_value),
                    message=f"{rule.error_message} ({target_name} is not numeric)",
                    severity=rule.severity,
                    suggested_fix=f"Ensure {target_name} contains a numeric value"
                ))
        
        except Exception as e:
            logger.error(f"Conditional sum error: {e}")
            raise
        
        return violations

    def _handle_tnt_bill_detection(self, rule: ValidationRule, df: pd.DataFrame, condition: str, workbook_data: Dict) -> List[ValidationViolation]:
        """Handle TNT bill detection: tnt_bill_detection:lookup_col:sum_col"""
        violations = []
        
        try:
            parts = condition.split(':')
            if len(parts) != 3:
                raise ValueError("TNT bill detection requires format: tnt_bill_detection:lookup_col:sum_col")
            
            lookup_col = parts[1].upper()
            sum_col = parts[2].upper()
            
            tnt_patterns = [
                "Subcon-Bill-to-TNT",
                "PDC-Bill-to-TNT", 
                "USI-Bill-to-TNT",
                "IFA-Bill-to-TNT"
            ]
            
            lookup_col_idx = self.base_interpreter.get_column_index_from_letter(lookup_col) - 1
            sum_col_idx = self.base_interpreter.get_column_index_from_letter(sum_col) - 1
            
            if lookup_col_idx >= len(df.columns) or sum_col_idx >= len(df.columns):
                return violations
            
            lookup_column_name = df.columns[lookup_col_idx]
            sum_column_name = df.columns[sum_col_idx]
            
            total_tnt_amount = 0
            tnt_entries = []
            
            for idx, row in df.iterrows():
                lookup_cell_value = row[lookup_column_name]
                sum_cell_value = row[sum_column_name]
                
                if pd.notna(lookup_cell_value):
                    lookup_str = str(lookup_cell_value).strip()
                    
                    for tnt_pattern in tnt_patterns:
                        if tnt_pattern.lower() in lookup_str.lower():
                            excel_row = idx + 2
                            
                            amount = 0
                            try:
                                if pd.notna(sum_cell_value) and str(sum_cell_value).strip() != '':
                                    amount = float(sum_cell_value)
                                    total_tnt_amount += amount
                            except (ValueError, TypeError):
                                pass
                            
                            tnt_entries.append({
                                'row': excel_row,
                                'pattern': tnt_pattern,
                                'lookup_value': lookup_str,
                                'amount': amount,
                                'sum_cell_addr': f"{sum_col}{excel_row}",
                                'lookup_cell_addr': f"{lookup_col}{excel_row}"
                            })
                            break
            
            # Create violations for each TNT entry found
            for entry in tnt_entries:
                violations.append(ValidationViolation(
                    rule_id=rule.rule_id,
                    sheet_name=rule.target_sheet,
                    cell_address=entry['lookup_cell_addr'],
                    row_index=entry['row'] - 2,
                    column_name=lookup_column_name,
                    value=entry['lookup_value'],
                    message=f"{rule.error_message} - {entry['pattern']} found (Amount: {self._format_number(entry['amount'])})",
                    severity=rule.severity,
                    suggested_fix=f"Review TNT billing entry and corresponding amount in {entry['sum_cell_addr']}"
                ))
                
                if entry['amount'] != 0:
                    violations.append(ValidationViolation(
                        rule_id=rule.rule_id,
                        sheet_name=rule.target_sheet,
                        cell_address=entry['sum_cell_addr'],
                        row_index=entry['row'] - 2,
                        column_name=sum_column_name,
                        value=self._format_number(entry['amount']),
                        message=f"{rule.error_message} - Amount for {entry['pattern']} (Total TNT: {self._format_number(total_tnt_amount)})",
                        severity=rule.severity,
                        suggested_fix=f"Verify amount {self._format_number(entry['amount'])} for {entry['pattern']} entry"
                    ))
        
        except Exception as e:
            logger.error(f"TNT bill detection error: {e}")
            raise
        
        return violations
    
    def _get_cell_display_name(self, cell_address: str) -> str:
        """Get user-friendly display name for cell addresses"""
        cell_names = {
            'H13': 'Recoverable OPE',
            'H14': 'Unrecoverable Expense',
            'H11': 'Contract Fee'
        }
        return cell_names.get(cell_address.upper(), cell_address)
    
    def _get_cell_value(self, df: pd.DataFrame, cell_address: str):
        """Get value from a specific cell address like H13"""
        try:
            cell_match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
            if not cell_match:
                return None
            
            col_letter = cell_match.group(1)
            excel_row = int(cell_match.group(2))
            
            col_idx = self.base_interpreter.get_column_index_from_letter(col_letter) - 1
            df_row_idx = excel_row - 2
            
            if col_idx >= 0 and col_idx < len(df.columns) and df_row_idx >= 0 and df_row_idx < len(df):
                return df.iloc[df_row_idx, col_idx]
            
            return None
        except Exception:
            return None

class RowValidatorAgent:
    """Enhanced validator that can handle both simple and smart rules"""
    
    def __init__(self, rule_interpreter: RuleInterpreterAgent, rule: ValidationRule):
        self.rule_interpreter = rule_interpreter
        self.rule = rule
        self.smart_interpreter = SmartRuleInterpreter(rule_interpreter)
        self.violations = []
    
    def validate(self, workbook_data: Dict[str, pd.DataFrame]) -> List[ValidationViolation]:
        """Enhanced validation that handles both simple and smart rules"""
        violations = []
        
        if self.rule.target_sheet not in workbook_data:
            return violations
        
        df = workbook_data[self.rule.target_sheet]
        condition = self.rule.condition.strip().lower()
        
        # Check if this is a smart rule
        if any(condition.startswith(smart_type) for smart_type in 
               ['gap_detection', 'conditional_comparison', 'conditional_sum', 'tnt_bill_detection']):
            violations = self.smart_interpreter.interpret_smart_rule(self.rule, df, workbook_data)
        else:
            # Handle as regular rule
            range_info = self.rule_interpreter.parse_range(self.rule.cell_column_range)
            
            if range_info['type'] == 'single_column':
                col_letter = range_info['column']
                col_index = self.rule_interpreter.get_column_index_from_letter(col_letter) - 1
                if col_index < len(df.columns):
                    column_name = df.columns[col_index]
                    violations.extend(self._validate_column(df, column_name, col_letter))
            
            elif range_info['type'] == 'column_range':
                start_col = range_info['start_col']
                end_col = range_info['end_col']
                start_idx = self.rule_interpreter.get_column_index_from_letter(start_col) - 1
                end_idx = self.rule_interpreter.get_column_index_from_letter(end_col) - 1
                for col_idx in range(start_idx, min(end_idx + 1, len(df.columns))):
                    column_name = df.columns[col_idx]
                    col_letter = self.rule_interpreter.get_column_letter_from_index(col_idx + 1)
                    violations.extend(self._validate_column(df, column_name, col_letter))
            
            elif range_info['type'] == 'cell_range':
                violations.extend(self._validate_cell_range(df, range_info))
            
            elif range_info['type'] == 'single_cell':
                violations.extend(self._validate_single_cell(df, range_info['cell']))
        
        return violations

    def _validate_column(self, df: pd.DataFrame, column_name: str, col_letter: str) -> List[ValidationViolation]:
        """Validate an entire column using this validator's rule"""
        violations = []
        
        for row_idx, value in enumerate(df[column_name]):
            if self.rule_interpreter.interpret_rule(self.rule, value):
                cell_address = f"{col_letter}{row_idx + 2}"
                
                violation = ValidationViolation(
                    rule_id=self.rule.rule_id,
                    sheet_name=self.rule.target_sheet,
                    cell_address=cell_address,
                    row_index=row_idx,
                    column_name=column_name,
                    value=value,
                    message=self.rule.error_message,
                    severity=self.rule.severity,
                    suggested_fix=self._suggest_fix(value)
                )
                violations.append(violation)
        
        return violations
    
    def _validate_cell_range(self, df: pd.DataFrame, range_info: Dict) -> List[ValidationViolation]:
        """Validate a specific cell range"""
        violations = []
        
        try:
            start_cell = range_info['start_cell']
            end_cell = range_info['end_cell']
            
            start_col_match = re.match(r'([A-Z]+)(\d+)', start_cell)
            end_col_match = re.match(r'([A-Z]+)(\d+)', end_cell)
            
            if not start_col_match or not end_col_match:
                return violations
            
            start_col_letter = start_col_match.group(1)
            start_row = int(start_col_match.group(2))
            end_col_letter = end_col_match.group(1)
            end_row = int(end_col_match.group(2))
            
            start_col_idx = self.rule_interpreter.get_column_index_from_letter(start_col_letter) - 1
            end_col_idx = self.rule_interpreter.get_column_index_from_letter(end_col_letter) - 1
            
            for col_idx in range(start_col_idx, min(end_col_idx + 1, len(df.columns))):
                if col_idx >= len(df.columns):
                    continue
                    
                column_name = df.columns[col_idx]
                col_letter = self.rule_interpreter.get_column_letter_from_index(col_idx + 1)
                
                for excel_row in range(start_row, min(end_row + 1, len(df) + 2)):
                    df_row_idx = excel_row - 2
                    
                    if df_row_idx >= 0 and df_row_idx < len(df):
                        value = df.iloc[df_row_idx, col_idx]
                        
                        if self.rule_interpreter.interpret_rule(self.rule, value):
                            cell_address = f"{col_letter}{excel_row}"
                            
                            violation = ValidationViolation(
                                rule_id=self.rule.rule_id,
                                sheet_name=self.rule.target_sheet,
                                cell_address=cell_address,
                                row_index=df_row_idx,
                                column_name=column_name,
                                value=value,
                                message=self.rule.error_message,
                                severity=self.rule.severity,
                                suggested_fix=self._suggest_fix(value)
                            )
                            violations.append(violation)
            
        except Exception as e:
            logger.error(f"Error validating cell range {range_info['range']}: {e}")
        
        return violations
    
    def _validate_single_cell(self, df: pd.DataFrame, cell_address: str) -> List[ValidationViolation]:
        """Validate a single cell"""
        violations = []
        
        try:
            cell_match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
            if not cell_match:
                return violations
            
            col_letter = cell_match.group(1)
            excel_row = int(cell_match.group(2))
            
            col_idx = self.rule_interpreter.get_column_index_from_letter(col_letter) - 1
            df_row_idx = excel_row - 2
            
            if col_idx >= 0 and col_idx < len(df.columns) and df_row_idx >= 0 and df_row_idx < len(df):
                column_name = df.columns[col_idx]
                value = df.iloc[df_row_idx, col_idx]
                
                if self.rule_interpreter.interpret_rule(self.rule, value):
                    violation = ValidationViolation(
                        rule_id=self.rule.rule_id,
                        sheet_name=self.rule.target_sheet,
                        cell_address=cell_address.upper(),
                        row_index=df_row_idx,
                        column_name=column_name,
                        value=value,
                        message=self.rule.error_message,
                        severity=self.rule.severity,
                        suggested_fix=self._suggest_fix(value)
                    )
                    violations.append(violation)
        
        except Exception as e:
            logger.error(f"Error validating single cell {cell_address}: {e}")
        
        return violations
    
    def _suggest_fix(self, value: Any) -> str:
        """Suggest fixes for common validation issues"""
        condition = self.rule.condition.lower()
        
        if '>0' in condition or '>= 0' in condition:
            try:
                suggested_value = abs(float(value)) if pd.notna(value) else 0
                formatted_value = f"{suggested_value:,.2f}".rstrip('0').rstrip('.')
                return f"Change to positive value: {formatted_value}"
            except:
                return "Change to positive value"
        elif 'date' in condition:
            return "Use appropriate date format"
        elif 'not_empty' in condition:
            return "Provide a value (cell is empty)"
        elif 'is_number' in condition:
            return "Provide a numeric value"
        else:
            return "Please review and correct the value"

class CorrectionAgent:
    """Agent that suggests and can auto-apply corrections"""
    
    def __init__(self):
        self.correction_strategies = {}
    
    def suggest_corrections(self, violations: List[ValidationViolation]) -> Dict[str, List[str]]:
        """Suggest corrections for all violations"""
        suggestions = {}
        
        for violation in violations:
            key = f"Row {violation.row_index + 2} - {violation.column_name}"
            if key not in suggestions:
                suggestions[key] = []
            
            suggestions[key].append(f"{violation.rule_id}: {violation.suggested_fix}")
        
        return suggestions

class SupervisorAgent:
    """Main orchestration agent that coordinates all other agents"""
    
    def __init__(self):
        self.rule_interpreter = RuleInterpreterAgent()
        self.validators = {}
        self.correction_agent = CorrectionAgent()
        
    def load_rules_from_excel_workbench(self, rules_file_content: bytes, rules_sheet_name: str = None) -> List[ValidationRule]:
        """Load validation rules from Excel and initialize per-rule validators"""
        try:
            excel_file = pd.ExcelFile(BytesIO(rules_file_content))
            available_sheets = excel_file.sheet_names
            
            if rules_sheet_name is None:
                for sheet in available_sheets:
                    if any(keyword in sheet.lower() for keyword in ['rule', 'config', 'validation', 'workbench']):
                        rules_sheet_name = sheet
                        break
                
                if rules_sheet_name is None:
                    rules_sheet_name = available_sheets[0]
            
            rules_df = pd.read_excel(BytesIO(rules_file_content), sheet_name=rules_sheet_name)
            
            if not isinstance(rules_df, pd.DataFrame):
                logger.error(f"Expected DataFrame but got {type(rules_df)}. Sheet '{rules_sheet_name}' may not exist.")
                return []
            
            rules = []
            for idx, row in rules_df.iterrows():
                try:
                    rule_id = self._get_column_value(row, ['Rule_ID', 'RuleID', 'ID', 'rule_id'], f'R{idx+1}')
                    target_sheet = self._get_column_value(row, ['Target Sheet', 'Sheet', 'Target_Sheet', 'target_sheet'], '')
                    cell_range = self._get_column_value(row, ['Cell/Column/Range', 'Range', 'Target', 'cell_column_range'], '')
                    error_message = self._get_column_value(row, ['Error message', 'Message', 'Error_Message', 'error_message'], '')
                    condition = self._get_column_value(row, ['Condition', 'Rule', 'Logic', 'condition'], '')
                    
                    if (str(target_sheet).strip() == '' or 
                        str(cell_range).strip() == '' or 
                        str(condition).strip() == '' or 
                        str(error_message).strip() == ''):
                        continue
                    
                    rule = ValidationRule(
                        rule_id=str(rule_id),
                        target_sheet=str(target_sheet),
                        cell_column_range=str(cell_range),
                        condition=str(condition),
                        error_message=str(error_message),
                        priority=idx + 1
                    )
                    rules.append(rule)
                    
                except Exception as e:
                    logger.warning(f"Could not parse rule in row {idx+1}: {e}")
                    continue
            
            if not rules:
                logger.error("No valid rules found in the rules file.")
                return []
            
            self.validators = {
                rule.rule_id: RowValidatorAgent(self.rule_interpreter, rule)
                for rule in rules
            }
            
            logger.info(f"Created {len(self.validators)} RowValidatorAgents for {len(rules)} rules")
            
            return rules
            
        except Exception as e:
            logger.error(f"Error loading workbench rules: {e}")
            return []

    def _get_column_value(self, row: pd.Series, possible_columns: List[str], default: Any = None) -> Any:
        """Get value from row using multiple possible column names"""
        row_columns_lower = [col.lower().replace('_', ' ').replace('/', '') for col in row.index]
        
        for col_name in possible_columns:
            search_name = col_name.lower().replace('_', ' ').replace('/', '')
            if search_name in row_columns_lower:
                actual_col = row.index[row_columns_lower.index(search_name)]
                value = row[actual_col]
                if pd.notna(value) and str(value).strip() != '' and str(value).strip().lower() not in ['nan', 'none', 'null']:
                    return value
        
        return default
        
    def process_workbook_with_workbench_rules(self, data_file_content: bytes, rules_file_content: bytes, rules_sheet_name: str = None) -> Dict:
        """Main processing pipeline with parallel rule validation"""
        rules = self.load_rules_from_excel_workbench(rules_file_content, rules_sheet_name)
        if not rules:
            return {}

        try:
            workbook_data = self._load_workbook(data_file_content)
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return {}

        all_violations = []
        
        try:
            max_workers = min(8, len(self.validators))
            logger.info(f"Running {len(self.validators)} validators with {max_workers} parallel workers")
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    rule_id: executor.submit(validator.validate, workbook_data)
                    for rule_id, validator in self.validators.items()
                }
                
                for rule_id, future in futures.items():
                    try:
                        rule_violations = future.result()
                        all_violations.extend(rule_violations)
                        logger.info(f"Rule {rule_id} completed: {len(rule_violations)} violations")
                    except Exception as e:
                        logger.error(f"Rule {rule_id} validation failed: {e}")
                        continue

        except Exception as e:
            logger.error(f"Parallel execution error: {e}")
            logger.warning("Falling back to sequential processing")
            all_violations = []
            for rule_id, validator in self.validators.items():
                try:
                    rule_violations = validator.validate(workbook_data)
                    all_violations.extend(rule_violations)
                    logger.info(f"Rule {rule_id} (sequential) completed: {len(rule_violations)} violations")
                except Exception as e:
                    logger.error(f"Sequential validation failed for rule {rule_id}: {e}")

        self.validation_results = self._compile_results(
            workbook_data=workbook_data,
            violations=all_violations,
            rules=rules
        )
        
        logger.info(f"Validation complete: {len(all_violations)} total violations found")
        return self.validation_results

    def _load_workbook(self, file_content: bytes) -> Dict[str, pd.DataFrame]:
        """Load all sheets from Excel workbook"""
        workbook_data = {}
        
        try:
            excel_file = pd.ExcelFile(BytesIO(file_content))
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name)
                    workbook_data[sheet_name] = df
                except Exception as e:
                    continue
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
        
        return workbook_data
    
    def _compile_results(self, workbook_data: Dict[str, pd.DataFrame], violations: List[ValidationViolation], rules: List[ValidationRule]) -> Dict:
        """Compile validation results"""
        violations_by_sheet = {}
        for violation in violations:
            sheet = violation.sheet_name
            if sheet not in violations_by_sheet:
                violations_by_sheet[sheet] = []
            violations_by_sheet[sheet].append(violation)
        
        total_cells_checked = sum(len(df) * len(df.columns) for df in workbook_data.values())
        
        results = {
            'workbook_data': workbook_data,
            'violations': violations,
            'violations_by_sheet': violations_by_sheet,
            'total_violations': len(violations),
            'total_cells_checked': total_cells_checked,
            'rules_applied': len(rules),
            'validation_status': 'PASSED' if len(violations) == 0 else 'FAILED',
            'summary': {
                'sheets_processed': len(workbook_data),
                'total_violations': len(violations),
                'affected_sheets': len(violations_by_sheet),
                'violation_rate': len(violations) / total_cells_checked if total_cells_checked > 0 else 0
            }
        }
        
        return results
    
    def generate_output_file(self, rules_file_content: bytes) -> bytes:
        """Generate the output Excel file with formatted headers and auto-adjusted columns"""
        if not self.validation_results:
            return None
        
        try:
            original_rules_df = pd.read_excel(BytesIO(rules_file_content))
            flagged_rules_df = original_rules_df.copy()
            
            added_columns = {
                'Status': 'OK',
                'violations_found': 0,
                'affected_cells': '',
                'sample_violations': ''
            }
            
            for col_name, default_value in added_columns.items():
                flagged_rules_df[col_name] = default_value
            
            violations_by_rule = {}
            for violation in self.validation_results['violations']:
                rule_id = violation.rule_id
                if rule_id not in violations_by_rule:
                    violations_by_rule[rule_id] = []
                violations_by_rule[rule_id].append(violation)
            
            for idx, row in flagged_rules_df.iterrows():
                rule_id = None
                for col in ['Rule_ID', 'rule_id', 'RuleID', 'ID']:
                    if col in row.index and pd.notna(row[col]):
                        rule_id = str(row[col])
                        break
                
                if rule_id and rule_id in violations_by_rule:
                    rule_violations = violations_by_rule[rule_id]
                    
                    flagged_rules_df.loc[idx, 'Status'] = 'NOT OK' if rule_violations else 'OK'
                    flagged_rules_df.loc[idx, 'violations_found'] = len(rule_violations)
                    
                    affected_cells = [f"{v.sheet_name}!{v.cell_address}" for v in rule_violations]
                    flagged_rules_df.loc[idx, 'affected_cells'] = ', '.join(affected_cells)
                    
                    sample_values = [str(v.value) for v in rule_violations]
                    flagged_rules_df.loc[idx, 'sample_violations'] = ', '.join(sample_values)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                flagged_rules_df.to_excel(writer, sheet_name='Validation_Results', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Validation_Results']
                
                from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
                
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                bold_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                for col_num, col_name in enumerate(flagged_rules_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = bold_font
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    
                    if col_name in added_columns:
                        cell.fill = green_fill
                    else:
                        cell.fill = white_fill
                    
                    try:
                        column_data = flagged_rules_df[col_name].astype(str)
                        max_data_len = column_data.map(len).max() if len(column_data) > 0 else 0
                        header_len = len(str(col_name))
                        max_len = max(max_data_len, header_len) + 2
                        
                        col_letter = cell.column_letter
                        worksheet.column_dimensions[col_letter].width = min(max_len, 50)
                    except Exception as e:
                        col_letter = cell.column_letter
                        worksheet.column_dimensions[col_letter].width = 15
                
                for row_num in range(2, len(flagged_rules_df) + 2):
                    for col_num in range(1, len(flagged_rules_df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        
                        if flagged_rules_df.columns[col_num-1] == 'Status':
                            if cell.value == 'NOT OK':
                                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                            elif cell.value == 'OK':
                                cell.fill = PatternFill(start_color="D6FFD6", end_color="D6FFD6", fill_type="solid")
                
                if self.validation_results['violations']:
                    detailed_violations = []
                    for violation in self.validation_results['violations']:
                        detailed_violations.append({
                            'Rule_ID': violation.rule_id,
                            'Sheet_Name': violation.sheet_name,
                            'Cell_Address': violation.cell_address,
                            'Column_Name': violation.column_name,
                            'Violating_Value': violation.value,
                            'Error_Message': violation.message,
                            'Suggested_Fix': violation.suggested_fix
                        })
                    
                    detailed_df = pd.DataFrame(detailed_violations)
                    detailed_df.to_excel(writer, sheet_name='Detailed_Violations', index=False)
                    
                    detailed_ws = writer.sheets['Detailed_Violations']
                    for col_num, col_name in enumerate(detailed_df.columns, 1):
                        cell = detailed_ws.cell(row=1, column=col_num)
                        cell.font = bold_font
                        cell.fill = white_fill
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        
                        try:
                            column_data = detailed_df[col_name].astype(str)
                            max_data_len = column_data.map(len).max() if len(column_data) > 0 else 0
                            header_len = len(str(col_name))
                            max_len = max(max_data_len, header_len) + 2
                            
                            col_letter = cell.column_letter
                            detailed_ws.column_dimensions[col_letter].width = min(max_len, 50)
                        except Exception as e:
                            col_letter = cell.column_letter
                            detailed_ws.column_dimensions[col_letter].width = 15
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating output file: {e}")
            return None