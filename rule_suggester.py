# rule_suggester.py
# AI-powered rule suggestion engine that analyzes Excel data patterns

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import re
from dataclasses import dataclass
import logging

logger = logging.getLogger(__name__)

@dataclass
class RuleSuggestion:
    """Represents an AI-suggested validation rule"""
    column_name: str
    rule_type: str
    suggested_condition: str
    confidence: float  # 0.0 to 1.0
    reasoning: str
    sample_values: List[str]
    priority: str  # 'high', 'medium', 'low'

class IntelligentRuleSuggester:
    """
    Analyzes Excel data and suggests appropriate validation rules using
    statistical analysis and pattern recognition
    """
    
    def __init__(self):
        self.patterns = {
            'email': r'^[\w\.-]+@[\w\.-]+\.\w+$',
            'phone_us': r'^\+?1?\s*\(?\d{3}\)?[\s\.-]?\d{3}[\s\.-]?\d{4}$',
            'ssn': r'^\d{3}-?\d{2}-?\d{4}$',
            'zip_us': r'^\d{5}(-\d{4})?$',
            'date_iso': r'^\d{4}-\d{2}-\d{2}$',
            'url': r'^https?://[\w\.-]+\.\w+',
            'uuid': r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        }
        
        self.financial_keywords = ['revenue', 'expense', 'cost', 'fee', 'price', 'amount', 'total', 'salary', 'payment']
        self.date_keywords = ['date', 'time', 'timestamp', 'created', 'updated', 'start', 'end']
        self.id_keywords = ['id', 'identifier', 'code', 'number', 'ref']
    
    def analyze_workbook(self, df: pd.DataFrame, sheet_name: str = None) -> List[RuleSuggestion]:
        """
        Analyze entire dataframe and return intelligent rule suggestions
        """
        all_suggestions = []
        
        for column in df.columns:
            column_suggestions = self.analyze_column(df[column], column, sheet_name)
            all_suggestions.extend(column_suggestions)
        
        # Sort by priority and confidence
        all_suggestions.sort(key=lambda x: (
            {'high': 3, 'medium': 2, 'low': 1}[x.priority],
            x.confidence
        ), reverse=True)
        
        return all_suggestions
    
    def analyze_column(self, column_data: pd.Series, column_name: str, sheet_name: str = None) -> List[RuleSuggestion]:
        """
        Deep analysis of a single column to suggest validation rules
        """
        suggestions = []
        
        # Skip if column is mostly empty
        non_null_count = column_data.notna().sum()
        if non_null_count < 3:
            return suggestions
        
        # Check for required field pattern (no nulls)
        if self._is_required_field(column_data):
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='required',
                suggested_condition='not_empty',
                confidence=0.95,
                reasoning=f'Column has no missing values across {len(column_data)} rows - likely required field',
                sample_values=self._get_sample_values(column_data, 3),
                priority='high'
            ))
        
        # Pattern-based detection
        pattern_suggestion = self._detect_pattern(column_data, column_name)
        if pattern_suggestion:
            suggestions.append(pattern_suggestion)
        
        # Numeric analysis
        if pd.api.types.is_numeric_dtype(column_data):
            numeric_suggestions = self._analyze_numeric_column(column_data, column_name)
            suggestions.extend(numeric_suggestions)
        
        # Date detection
        date_suggestion = self._detect_date_column(column_data, column_name)
        if date_suggestion:
            suggestions.append(date_suggestion)
        
        # Text pattern analysis
        if column_data.dtype == 'object':
            text_suggestions = self._analyze_text_column(column_data, column_name)
            suggestions.extend(text_suggestions)
        
        return suggestions
    
    def _is_required_field(self, column_data: pd.Series) -> bool:
        """Check if column appears to be required (no nulls)"""
        return column_data.notna().mean() == 1.0 and len(column_data) > 0
    
    def _detect_pattern(self, column_data: pd.Series, column_name: str) -> Optional[RuleSuggestion]:
        """Detect common data patterns (email, phone, SSN, etc.)"""
        
        # Get sample of non-null values
        sample = column_data.dropna().astype(str).head(100)
        if len(sample) == 0:
            return None
        
        # Test each pattern
        for pattern_name, pattern_regex in self.patterns.items():
            try:
                matches = sample.str.match(pattern_regex, case=False)
                match_rate = matches.sum() / len(sample)
                
                if match_rate > 0.75:  # 75% match threshold
                    return RuleSuggestion(
                        column_name=column_name,
                        rule_type=pattern_name,
                        suggested_condition=f'regex:{pattern_regex}',
                        confidence=min(match_rate, 0.95),
                        reasoning=f'Detected {pattern_name} pattern in {match_rate*100:.1f}% of values',
                        sample_values=self._get_sample_values(column_data, 3),
                        priority='high' if pattern_name in ['email', 'ssn', 'phone_us'] else 'medium'
                    )
            except Exception as e:
                continue
        
        return None
    
    def _analyze_numeric_column(self, column_data: pd.Series, column_name: str) -> List[RuleSuggestion]:
        """Analyze numeric columns for range and outlier detection"""
        suggestions = []
        
        # Remove nulls for analysis
        clean_data = column_data.dropna()
        if len(clean_data) < 3:
            return suggestions
        
        # Statistical outlier detection using IQR
        q1 = clean_data.quantile(0.25)
        q3 = clean_data.quantile(0.75)
        iqr = q3 - q1
        
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        
        # Check if data is naturally bounded (e.g., all positive)
        min_val = clean_data.min()
        max_val = clean_data.max()
        mean_val = clean_data.mean()
        
        # Detect negative values in financial columns (potential errors)
        if any(keyword in column_name.lower() for keyword in self.financial_keywords):
            if (clean_data < 0).any():
                negative_pct = (clean_data < 0).sum() / len(clean_data) * 100
                
                if negative_pct < 5:  # Less than 5% negative
                    suggestions.append(RuleSuggestion(
                        column_name=column_name,
                        rule_type='positive_financial',
                        suggested_condition='>= 0',
                        confidence=0.85,
                        reasoning=f'Financial column with {negative_pct:.1f}% negative values - may be errors',
                        sample_values=self._get_sample_values(clean_data[clean_data < 0], 3),
                        priority='high'
                    ))
        
        # Suggest range based on IQR (for outlier detection)
        outliers = ((clean_data < lower_bound) | (clean_data > upper_bound)).sum()
        outlier_pct = outliers / len(clean_data) * 100
        
        if outlier_pct > 2 and outlier_pct < 20:  # Between 2-20% outliers
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='range_validation',
                suggested_condition=f'range:{lower_bound:.2f}:{upper_bound:.2f}',
                confidence=0.70,
                reasoning=f'Statistical outlier detection: {outliers} values ({outlier_pct:.1f}%) outside IQR bounds',
                sample_values=self._get_sample_values(clean_data[(clean_data < lower_bound) | (clean_data > upper_bound)], 3),
                priority='medium'
            ))
        
        # Detect percentage columns (values between 0-1 or 0-100)
        if min_val >= 0 and max_val <= 1:
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='percentage',
                suggested_condition='range:0:1',
                confidence=0.90,
                reasoning='Values appear to be percentages (0-1 range)',
                sample_values=self._get_sample_values(clean_data, 3),
                priority='medium'
            ))
        elif min_val >= 0 and max_val <= 100 and 'percent' in column_name.lower():
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='percentage',
                suggested_condition='range:0:100',
                confidence=0.85,
                reasoning='Values appear to be percentages (0-100 range)',
                sample_values=self._get_sample_values(clean_data, 3),
                priority='medium'
            ))
        
        return suggestions
    
    def _detect_date_column(self, column_data: pd.Series, column_name: str) -> Optional[RuleSuggestion]:
        """Detect date/datetime columns"""
        
        # Check column name first
        if not any(keyword in column_name.lower() for keyword in self.date_keywords):
            return None
        
        # Try to parse as dates
        sample = column_data.dropna().head(50)
        if len(sample) == 0:
            return None
        
        try:
            parsed = pd.to_datetime(sample, errors='coerce')
            valid_dates = parsed.notna().sum()
            success_rate = valid_dates / len(sample)
            
            if success_rate > 0.80:
                return RuleSuggestion(
                    column_name=column_name,
                    rule_type='date_validation',
                    suggested_condition='is_date',
                    confidence=min(success_rate, 0.95),
                    reasoning=f'Successfully parsed {success_rate*100:.1f}% of values as dates',
                    sample_values=self._get_sample_values(column_data, 3),
                    priority='high'
                )
        except Exception:
            pass
        
        return None
    
    def _analyze_text_column(self, column_data: pd.Series, column_name: str) -> List[RuleSuggestion]:
        """Analyze text columns for patterns"""
        suggestions = []
        
        clean_data = column_data.dropna().astype(str)
        if len(clean_data) < 3:
            return suggestions
        
        # Check for consistent length (like SKU codes, ID numbers)
        lengths = clean_data.str.len()
        if lengths.std() == 0 and lengths.mean() > 3:  # All same length
            avg_length = int(lengths.mean())
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='fixed_length',
                suggested_condition=f'length:{avg_length}',
                confidence=0.90,
                reasoning=f'All values have consistent length ({avg_length} characters) - likely structured code',
                sample_values=self._get_sample_values(clean_data, 3),
                priority='medium'
            ))
        
        # Check for categorical data (low cardinality)
        unique_ratio = len(clean_data.unique()) / len(clean_data)
        if unique_ratio < 0.10 and len(clean_data.unique()) < 20:  # Less than 10% unique
            unique_values = clean_data.unique()[:5]
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='categorical',
                suggested_condition=f'in:{",".join(map(str, unique_values))}',
                confidence=0.85,
                reasoning=f'Low cardinality ({len(clean_data.unique())} unique values) - appears categorical',
                sample_values=list(map(str, unique_values)),
                priority='medium'
            ))
        
        # Check for uppercase/lowercase consistency
        if (clean_data.str.isupper()).mean() > 0.95:
            suggestions.append(RuleSuggestion(
                column_name=column_name,
                rule_type='format_uppercase',
                suggested_condition='is_uppercase',
                confidence=0.80,
                reasoning='95%+ values are uppercase - likely formatting standard',
                sample_values=self._get_sample_values(clean_data, 3),
                priority='low'
            ))
        
        return suggestions
    
    def _get_sample_values(self, series: pd.Series, n: int = 3) -> List[str]:
        """Get sample values from series for display"""
        sample = series.dropna().head(n)
        return [str(val)[:50] for val in sample]  # Limit to 50 chars
    
    def suggest_cross_field_rules(self, df: pd.DataFrame) -> List[RuleSuggestion]:
        """
        Suggest validation rules that involve multiple columns
        (e.g., start_date < end_date, sum validations)
        """
        suggestions = []
        
        # Find potential date pairs
        date_columns = [col for col in df.columns if any(kw in col.lower() for kw in self.date_keywords)]
        
        for i, col1 in enumerate(date_columns):
            for col2 in date_columns[i+1:]:
                if ('start' in col1.lower() and 'end' in col2.lower()) or \
                   ('begin' in col1.lower() and 'end' in col2.lower()) or \
                   ('from' in col1.lower() and 'to' in col2.lower()):
                    
                    suggestions.append(RuleSuggestion(
                        column_name=f"{col1} vs {col2}",
                        rule_type='date_range',
                        suggested_condition=f'cross_field:{col1}:{col2}:less_than',
                        confidence=0.85,
                        reasoning=f'Detected date range pattern: {col1} should be before {col2}',
                        sample_values=[],
                        priority='high'
                    ))
        
        # Find potential sum relationships
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if 'total' in ' '.join(numeric_cols).lower():
            total_cols = [col for col in numeric_cols if 'total' in col.lower()]
            for total_col in total_cols:
                suggestions.append(RuleSuggestion(
                    column_name=total_col,
                    rule_type='sum_validation',
                    suggested_condition=f'validate_sum:{total_col}',
                    confidence=0.70,
                    reasoning=f'Column "{total_col}" may be sum of other columns - consider adding sum validation',
                    sample_values=[],
                    priority='medium'
                ))
        
        return suggestions
    
    def generate_rules_excel(self, suggestions: List[RuleSuggestion], output_path: str):
        """
        Generate an Excel file with suggested rules in the proper format
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Suggested_Rules"
        
        # Headers
        headers = ["Rule_ID", "Target Sheet", "Cell/Column/Range", "Condition", 
                   "Error message", "Active", "Confidence", "Reasoning"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add suggestions
        for idx, suggestion in enumerate(suggestions, 1):
            ws.append([
                f"AI_{idx}",
                "Sheet1",  # User needs to update
                suggestion.column_name,
                suggestion.suggested_condition,
                suggestion.reasoning,
                "YES",
                f"{suggestion.confidence*100:.0f}%",
                suggestion.reasoning
            ])
            
            # Color code by priority
            row_idx = idx + 1
            if suggestion.priority == 'high':
                fill_color = "C6EFCE"  # Light green
            elif suggestion.priority == 'medium':
                fill_color = "FFEB9C"  # Light yellow
            else:
                fill_color = "FFC7CE"  # Light red
            
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Generated suggested rules file: {output_path}")
