from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import os
import tempfile
import pandas as pd
from datetime import datetime
from io import BytesIO
import logging
import json

# Import all validation logic from validation_logic.py
from validation_logic import (
    ValidationRule,
    ValidationViolation,
    SupervisorAgent,
    RuleInterpreterAgent,
    SmartRuleInterpreter,
    RowValidatorAgent,
    CorrectionAgent
)

# Import AI rule suggester
from rule_suggester import IntelligentRuleSuggester

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

def generate_smart_excel_template() -> bytes:
    """Generate ultra-simple, user-friendly template"""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.comments import Comment
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ===== MAIN RULES SHEET (SUPER SIMPLE) =====
    rule_sheet = wb.create_sheet("Rules")
    
    # Simple headers
    headers = [
        "Rule_ID", "Target Sheet", "Cell/Column/Range", "Error message", "Condition", "Active"
    ]
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = rule_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # YOUR 4 SPECIFIC RULES
    example_rules = [
        ["R1", "Actuals (Weeks)", "D:JC:24:31", "Missing week entry between project start and end", "gap_detection:D:JC:24:31", "YES"],
        ["R2", "Forecast", "H13:H14", "Recoverable OPE cannot exceed Unrecoverable Expense or be non-zero when Unrecoverable Expense is zero", "conditional_comparison:H13:H14:greater_or_zero", "YES"],
        ["R3", "Forecast", "G:K:H11", "DC-Bill-to-Client entries must sum to Contract Fee total", "conditional_sum:G:DC-Bill-to-Client:K:H11", "YES"],
        ["R4", "Forecast", "G:K", "TNT billing entries detected - review amounts and totals", "tnt_bill_detection:G:K", "YES"],
    ]
    
    # Add data
    for row, rule_data in enumerate(example_rules, 2):
        for col, value in enumerate(rule_data, 1):
            cell = rule_sheet.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if row in [2, 3, 4, 5]:
                cell.font = Font(bold=True)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# API Routes

@app.route('/')
def index():
    """Serve the main HTML page"""
    return render_template('index.html')

@app.route('/api/download-template', methods=['GET'])
def download_template():
    """Generate and download rule template"""
    try:
        template_bytes = generate_smart_excel_template()
        
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(template_bytes)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='Rule_Template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        return jsonify({'error': 'Failed to generate template'}), 500

@app.route('/api/upload-files', methods=['POST'])
def upload_files():
    """Handle file uploads and return sheet information"""
    try:
        if 'data_file' not in request.files or 'rules_file' not in request.files:
            return jsonify({'error': 'Both data file and rules file are required'}), 400
        
        data_file = request.files['data_file']
        rules_file = request.files['rules_file']
        
        if data_file.filename == '' or rules_file.filename == '':
            return jsonify({'error': 'No files selected'}), 400
        
        # Save files temporarily
        data_filename = os.path.join(app.config['UPLOAD_FOLDER'], f"data_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{data_file.filename}")
        rules_filename = os.path.join(app.config['UPLOAD_FOLDER'], f"rules_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{rules_file.filename}")
        
        data_file.save(data_filename)
        rules_file.save(rules_filename)
        
        # Get sheet names from rules file
        try:
            excel_file = pd.ExcelFile(rules_filename)
            available_sheets = excel_file.sheet_names
        except Exception as e:
            return jsonify({'error': f'Error reading rules file: {e}'}), 400
        
        return jsonify({
            'success': True,
            'data_filename': data_filename,
            'rules_filename': rules_filename,
            'available_sheets': available_sheets,
            'data_file_original': data_file.filename,
            'rules_file_original': rules_file.filename
        })
        
    except Exception as e:
        logger.error(f"Error uploading files: {e}")
        return jsonify({'error': f'Upload failed: {e}'}), 500

@app.route('/api/validate', methods=['POST'])
def validate_files():
    """Process validation"""
    try:
        data = request.get_json()
        data_filename = data.get('data_filename')
        rules_filename = data.get('rules_filename')
        rules_sheet_name = data.get('rules_sheet_name')
        
        if not data_filename or not rules_filename:
            return jsonify({'error': 'File paths not provided'}), 400
        
        # Read file contents
        with open(data_filename, 'rb') as f:
            data_content = f.read()
        
        with open(rules_filename, 'rb') as f:
            rules_content = f.read()
        
        # Initialize supervisor agent and process
        supervisor = SupervisorAgent()  # You'll need to include this class
        results = supervisor.process_workbook_with_workbench_rules(
            data_content, 
            rules_content, 
            rules_sheet_name
        )
        
        if not results:
            return jsonify({'error': 'Validation failed - no results generated'}), 500
        
        # Convert results to JSON-serializable format
        violations_data = []
        for violation in results['violations']:
            violations_data.append({
                'rule_id': violation.rule_id,
                'sheet_name': violation.sheet_name,
                'cell_address': violation.cell_address,
                'row_index': violation.row_index,
                'column_name': violation.column_name,
                'value': str(violation.value),
                'message': violation.message,
                'severity': violation.severity,
                'suggested_fix': violation.suggested_fix
            })
        
        # Store results for download (you might want to use a database or cache)
        session_id = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        results_filename = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_results.json")
        
        results_data = {
            'violations': violations_data,
            'total_violations': results['total_violations'],
            'total_cells_checked': results['total_cells_checked'],
            'rules_applied': results['rules_applied'],
            'validation_status': results['validation_status'],
            'summary': results['summary'],
            'rules_content': rules_content.hex()  # Store as hex for JSON serialization
        }
        
        with open(results_filename, 'w') as f:
            json.dump(results_data, f)
        
        # Clean up temporary files
        try:
            os.remove(data_filename)
            os.remove(rules_filename)
        except:
            pass
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'total_violations': results['total_violations'],
            'validation_status': results['validation_status'],
            'summary': results['summary'],
            'violations': violations_data[:10]  # Return first 10 violations for preview
        })
        
    except Exception as e:
        logger.error(f"Error during validation: {e}")
        return jsonify({'error': f'Validation failed: {e}'}), 500

@app.route('/api/download-report/<session_id>', methods=['GET'])
def download_report(session_id):
    """Generate and download validation report"""
    try:
        results_filename = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_results.json")
        
        if not os.path.exists(results_filename):
            return jsonify({'error': 'Results not found'}), 404
        
        # Load results
        with open(results_filename, 'r') as f:
            results_data = json.load(f)
        
        # Convert back to original format
        rules_content = bytes.fromhex(results_data['rules_content'])
        
        # Generate report (you'll need to adapt your existing report generation)
        supervisor = SupervisorAgent()
        
        # Reconstruct validation results
        violations = []
        for v_data in results_data['violations']:
            violation = ValidationViolation(
                rule_id=v_data['rule_id'],
                sheet_name=v_data['sheet_name'],
                cell_address=v_data['cell_address'],
                row_index=v_data['row_index'],
                column_name=v_data['column_name'],
                value=v_data['value'],
                message=v_data['message'],
                severity=v_data['severity'],
                suggested_fix=v_data['suggested_fix']
            )
            violations.append(violation)
        
        supervisor.validation_results = {
            'violations': violations,
            'total_violations': results_data['total_violations'],
            'summary': results_data['summary']
        }
        
        output_file = supervisor.generate_output_file(rules_content)
        
        if output_file:
            # Create temporary file for download
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(output_file)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name=f'validation_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({'error': 'Failed to generate report'}), 500
            
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        return jsonify({'error': f'Report generation failed: {e}'}), 500

@app.route('/api/suggest-rules', methods=['POST'])
def suggest_rules():
    """AI-powered rule suggestion endpoint"""
    try:
        if 'data_file' not in request.files:
            return jsonify({'error': 'No data file provided'}), 400
        
        data_file = request.files['data_file']
        
        if data_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read the Excel file
        df = pd.read_excel(data_file)
        
        # Initialize suggester
        suggester = IntelligentRuleSuggester()
        
        # Generate suggestions
        suggestions = suggester.analyze_workbook(df)
        cross_field_suggestions = suggester.suggest_cross_field_rules(df)
        all_suggestions = suggestions + cross_field_suggestions
        
        # Limit to top 20 suggestions
        top_suggestions = all_suggestions[:20]
        
        # Generate downloadable Excel file with suggestions
        output_filename = os.path.join(app.config['UPLOAD_FOLDER'], 
                                      f"suggested_rules_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        suggester.generate_rules_excel(top_suggestions, output_filename)
        
        # Convert suggestions to JSON
        suggestions_json = []
        for s in top_suggestions:
            suggestions_json.append({
                'column_name': s.column_name,
                'rule_type': s.rule_type,
                'suggested_condition': s.suggested_condition,
                'confidence': round(s.confidence * 100, 1),
                'reasoning': s.reasoning,
                'sample_values': s.sample_values,
                'priority': s.priority
            })
        
        return jsonify({
            'success': True,
            'total_suggestions': len(top_suggestions),
            'suggestions': suggestions_json,
            'download_file': output_filename
        })
        
    except Exception as e:
        logger.error(f"Error suggesting rules: {e}")
        return jsonify({'error': f'Rule suggestion failed: {e}'}), 500

@app.route('/api/download-suggested-rules/<filename>', methods=['GET'])
def download_suggested_rules(filename):
    """Download the AI-generated rules file"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(filename))
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f'AI_suggested_rules_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error downloading suggested rules: {e}")
        return jsonify({'error': f'Download failed: {e}'}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Simple health check endpoint"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)